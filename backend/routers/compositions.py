import csv
import io
import json
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from backend.database import get_db
from backend.schemas.composition import CompositionResponse, HoldingItem
from backend.schemas.comparison import (
    CompareResponse, CompareSummary,
    AddedComponent, RemovedComponent, WeightChange,
    ChangesResponse, ChangeEvent,
)

router = APIRouter(prefix="/api/etfs", tags=["compositions"])

PAGE_SIZE = 100


@router.get("/{ticker}/compositions/{date}", response_model=CompositionResponse)
def get_composition(
    ticker: str,
    date: str,
    limit: int = Query(default=PAGE_SIZE, ge=1, le=5000),
    offset: int = Query(default=0, ge=0),
):
    ticker = ticker.upper()
    with get_db() as conn:
        meta = conn.execute(
            "SELECT nav, shares_outstanding, total_net_assets "
            "FROM equity_etf_compositions WHERE etf_ticker=? AND composition_date=? LIMIT 1",
            (ticker, date),
        ).fetchone()
        if not meta:
            raise HTTPException(status_code=404, detail=f"No composition for {ticker} on {date}")

        total_count = conn.execute(
            "SELECT COUNT(*) FROM equity_etf_compositions WHERE etf_ticker=? AND composition_date=?",
            (ticker, date),
        ).fetchone()[0]

        rows = conn.execute(
            "SELECT component_identifier, component_name, component_ticker, "
            "component_weight, component_sector, component_shares, component_currency, component_sedol "
            "FROM equity_etf_compositions WHERE etf_ticker=? AND composition_date=? "
            "ORDER BY component_weight DESC LIMIT ? OFFSET ?",
            (ticker, date, limit, offset),
        ).fetchall()

    return CompositionResponse(
        ticker=ticker,
        composition_date=date,
        nav=meta["nav"],
        shares_outstanding=meta["shares_outstanding"],
        total_net_assets=meta["total_net_assets"],
        holdings_count=total_count,
        holdings=[HoldingItem(**dict(r)) for r in rows],
        holdings_truncated=(offset + limit < total_count),
    )


@router.get("/{ticker}/compare", response_model=CompareResponse)
def compare_compositions(
    ticker: str,
    date1: str = Query(..., description="Earlier date YYYY-MM-DD"),
    date2: str = Query(..., description="Later date YYYY-MM-DD"),
):
    ticker = ticker.upper()
    if date1 == date2:
        raise HTTPException(status_code=400, detail="date1 and date2 must be different")

    with get_db() as conn:
        # Validate both dates exist
        for d in (date1, date2):
            exists = conn.execute(
                "SELECT 1 FROM equity_etf_compositions WHERE etf_ticker=? AND composition_date=? LIMIT 1",
                (ticker, d),
            ).fetchone()
            if not exists:
                raise HTTPException(status_code=404, detail=f"No composition for {ticker} on {d}")

        # Additions: in date2, not in date1
        added_rows = conn.execute(
            """
            SELECT c2.component_identifier, c2.component_name, c2.component_ticker,
                   c2.component_weight AS weight_new, c2.component_shares AS shares_new,
                   c2.component_sector, c2.component_currency
            FROM equity_etf_compositions c2
            LEFT JOIN equity_etf_compositions c1
                ON  c1.etf_ticker           = c2.etf_ticker
                AND c1.component_identifier = c2.component_identifier
                AND c1.composition_date     = ?
            WHERE c2.etf_ticker       = ?
              AND c2.composition_date = ?
              AND c1.component_identifier IS NULL
            ORDER BY c2.component_weight DESC
            """,
            (date1, ticker, date2),
        ).fetchall()

        # Removals: in date1, not in date2
        removed_rows = conn.execute(
            """
            SELECT c1.component_identifier, c1.component_name, c1.component_ticker,
                   c1.component_weight AS weight_old, c1.component_shares AS shares_old,
                   c1.component_sector, c1.component_currency
            FROM equity_etf_compositions c1
            LEFT JOIN equity_etf_compositions c2
                ON  c2.etf_ticker           = c1.etf_ticker
                AND c2.component_identifier = c1.component_identifier
                AND c2.composition_date     = ?
            WHERE c1.etf_ticker       = ?
              AND c1.composition_date = ?
              AND c2.component_identifier IS NULL
            ORDER BY c1.component_weight DESC
            """,
            (date2, ticker, date1),
        ).fetchall()

        # Weight changes: present in both, shares count actually changed
        changed_rows = conn.execute(
            """
            SELECT c1.component_identifier, c1.component_name, c1.component_ticker,
                   c1.component_weight AS weight_old, c2.component_weight AS weight_new,
                   (c2.component_weight - c1.component_weight) AS weight_delta,
                   c1.component_sector,
                   c1.component_shares AS shares_old, c2.component_shares AS shares_new,
                   (c2.component_shares - c1.component_shares) AS shares_delta
            FROM equity_etf_compositions c1
            JOIN equity_etf_compositions c2
                ON  c2.etf_ticker           = c1.etf_ticker
                AND c2.component_identifier = c1.component_identifier
                AND c2.composition_date     = ?
            WHERE c1.etf_ticker       = ?
              AND c1.composition_date = ?
              AND c1.component_shares IS NOT NULL AND c2.component_shares IS NOT NULL
              AND c1.component_shares != c2.component_shares
            ORDER BY ABS(c2.component_shares - c1.component_shares) DESC
            """,
            (date2, ticker, date1),
        ).fetchall()

    added = [AddedComponent(**dict(r)) for r in added_rows]
    removed = [RemovedComponent(**dict(r)) for r in removed_rows]
    weight_changes = [WeightChange(**dict(r)) for r in changed_rows]

    return CompareResponse(
        ticker=ticker,
        date1=date1,
        date2=date2,
        summary=CompareSummary(
            added_count=len(added),
            removed_count=len(removed),
            weight_changes_count=len(weight_changes),
        ),
        added=added,
        removed=removed,
        weight_changes=weight_changes,
    )


def _compute_changes_for_pair(conn, ticker: str, date_from: str, date_to: str) -> list[dict]:
    """Compute and return change rows for a consecutive date pair. Does not commit."""
    changes = []

    added_rows = conn.execute(
        """
        SELECT c2.component_identifier, c2.component_name, c2.component_ticker,
               c2.component_sector, c2.component_weight AS weight_new,
               c2.component_shares AS shares_new
        FROM equity_etf_compositions c2
        LEFT JOIN equity_etf_compositions c1
            ON c1.etf_ticker=c2.etf_ticker AND c1.component_identifier=c2.component_identifier AND c1.composition_date=?
        WHERE c2.etf_ticker=? AND c2.composition_date=? AND c1.component_identifier IS NULL
        """,
        (date_from, ticker, date_to),
    ).fetchall()
    for r in added_rows:
        changes.append({
            "etf_ticker": ticker, "date_from": date_from, "date_to": date_to,
            "change_type": "added",
            "component_identifier": r["component_identifier"],
            "component_name": r["component_name"],
            "component_ticker": r["component_ticker"],
            "component_sector": r["component_sector"],
            "weight_old": None, "weight_new": r["weight_new"], "weight_delta": None,
            "shares_old": None, "shares_new": r["shares_new"],
        })

    removed_rows = conn.execute(
        """
        SELECT c1.component_identifier, c1.component_name, c1.component_ticker,
               c1.component_sector, c1.component_weight AS weight_old,
               c1.component_shares AS shares_old
        FROM equity_etf_compositions c1
        LEFT JOIN equity_etf_compositions c2
            ON c2.etf_ticker=c1.etf_ticker AND c2.component_identifier=c1.component_identifier AND c2.composition_date=?
        WHERE c1.etf_ticker=? AND c1.composition_date=? AND c2.component_identifier IS NULL
        """,
        (date_to, ticker, date_from),
    ).fetchall()
    for r in removed_rows:
        changes.append({
            "etf_ticker": ticker, "date_from": date_from, "date_to": date_to,
            "change_type": "removed",
            "component_identifier": r["component_identifier"],
            "component_name": r["component_name"],
            "component_ticker": r["component_ticker"],
            "component_sector": r["component_sector"],
            "weight_old": r["weight_old"], "weight_new": None, "weight_delta": None,
            "shares_old": r["shares_old"], "shares_new": None,
        })

    changed_rows = conn.execute(
        """
        SELECT c1.component_identifier, c1.component_name, c1.component_ticker,
               c1.component_sector, c1.component_weight AS weight_old,
               c2.component_weight AS weight_new,
               (c2.component_weight - c1.component_weight) AS weight_delta,
               c1.component_shares AS shares_old,
               c2.component_shares AS shares_new
        FROM equity_etf_compositions c1
        JOIN equity_etf_compositions c2
            ON c2.etf_ticker=c1.etf_ticker AND c2.component_identifier=c1.component_identifier AND c2.composition_date=?
        WHERE c1.etf_ticker=? AND c1.composition_date=?
          AND c1.component_shares IS NOT NULL AND c2.component_shares IS NOT NULL
          AND c1.component_shares != c2.component_shares
        """,
        (date_to, ticker, date_from),
    ).fetchall()
    for r in changed_rows:
        changes.append({
            "etf_ticker": ticker, "date_from": date_from, "date_to": date_to,
            "change_type": "weight_change",
            "component_identifier": r["component_identifier"],
            "component_name": r["component_name"],
            "component_ticker": r["component_ticker"],
            "component_sector": r["component_sector"],
            "weight_old": r["weight_old"],
            "weight_new": r["weight_new"],
            "weight_delta": r["weight_delta"],
            "shares_old": r["shares_old"],
            "shares_new": r["shares_new"],
        })

    return changes


def _build_xlsx(ticker: str, date: str, info: dict, comp_meta: dict, holdings: list[dict]) -> StreamingResponse:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    center = Alignment(horizontal="center")

    # Sheet 1: ETF Info
    ws1 = wb.active
    ws1.title = "ETF Info"
    ws1.append(["Field", "Value"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in [
        ("Ticker", ticker),
        ("Name", info.get("name") or ""),
        ("Composition Date", date),
        ("NAV", info.get("nav") or ""),
        ("AUM", info.get("aum") or ""),
        ("Expense Ratio", info.get("gross_expense_ratio") or ""),
        ("Domicile", info.get("domicile") or ""),
        ("Shares Outstanding", comp_meta.get("shares_outstanding") or ""),
        ("Total Net Assets", comp_meta.get("total_net_assets") or ""),
    ]:
        ws1.append(list(row))
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 42

    # Sheet 2: Holdings
    ws2 = wb.create_sheet("Holdings")
    headers = ["Ticker", "Name", "SEDOL", "Identifier", "Weight (%)", "Shares", "Sector", "Currency"]
    ws2.append(headers)
    for i, _ in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for h in holdings:
        ws2.append([
            h.get("component_ticker") or "",
            h.get("component_name") or "",
            h.get("component_sedol") or "",
            h.get("component_identifier") or "",
            h.get("component_weight"),
            h.get("component_shares"),
            h.get("component_sector") or "",
            h.get("component_currency") or "",
        ])

    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0.0000"

    for i, w in enumerate([10, 42, 12, 22, 12, 16, 28, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{ticker}_{date}_composition.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_csv(ticker: str, date: str, info: dict, comp_meta: dict, holdings: list[dict]) -> StreamingResponse:
    buf = io.StringIO()
    buf.write('\ufeff')  # UTF-8 BOM so Excel auto-detects encoding
    writer = csv.writer(buf)

    # Metadata section
    writer.writerow(["Field", "Value"])
    for row in [
        ("Ticker", ticker),
        ("Name", info.get("name") or ""),
        ("Composition Date", date),
        ("NAV", info.get("nav") or ""),
        ("AUM", info.get("aum") or ""),
        ("Expense Ratio", info.get("gross_expense_ratio") or ""),
        ("Domicile", info.get("domicile") or ""),
        ("Shares Outstanding", comp_meta.get("shares_outstanding") or ""),
        ("Total Net Assets", comp_meta.get("total_net_assets") or ""),
    ]:
        writer.writerow(list(row))

    writer.writerow([])  # blank separator

    # Holdings section
    writer.writerow(["Ticker", "Name", "SEDOL", "Identifier", "Weight (%)", "Shares", "Sector", "Currency"])
    for h in holdings:
        writer.writerow([
            h.get("component_ticker") or "",
            h.get("component_name") or "",
            h.get("component_sedol") or "",
            h.get("component_identifier") or "",
            h.get("component_weight"),
            h.get("component_shares"),
            h.get("component_sector") or "",
            h.get("component_currency") or "",
        ])

    buf.seek(0)
    filename = f"{ticker}_{date}_composition.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_json(ticker: str, date: str, info: dict, comp_meta: dict, holdings: list[dict]) -> StreamingResponse:
    payload = {
        "ticker": ticker,
        "name": info.get("name"),
        "composition_date": date,
        "nav": info.get("nav"),
        "aum": info.get("aum"),
        "expense_ratio": info.get("gross_expense_ratio"),
        "domicile": info.get("domicile"),
        "shares_outstanding": comp_meta.get("shares_outstanding"),
        "total_net_assets": comp_meta.get("total_net_assets"),
        "holdings_count": len(holdings),
        "holdings": [
            {
                "ticker": h.get("component_ticker"),
                "name": h.get("component_name"),
                "sedol": h.get("component_sedol"),
                "identifier": h.get("component_identifier"),
                "weight_pct": h.get("component_weight"),
                "shares": h.get("component_shares"),
                "sector": h.get("component_sector"),
                "currency": h.get("component_currency"),
            }
            for h in holdings
        ],
    }
    filename = f"{ticker}_{date}_composition.json"
    return StreamingResponse(
        iter([json.dumps(payload, indent=2, ensure_ascii=False)]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{ticker}/compositions/{date}/download")
def download_composition(
    ticker: str,
    date: str,
    format: str = Query(default="xlsx"),
):
    ticker = ticker.upper()
    with get_db() as conn:
        info = conn.execute(
            "SELECT * FROM equity_etf_info WHERE ticker=?", (ticker,)
        ).fetchone()
        if not info:
            raise HTTPException(status_code=404, detail=f"ETF '{ticker}' not found")

        comp_meta = conn.execute(
            "SELECT nav, shares_outstanding, total_net_assets "
            "FROM equity_etf_compositions WHERE etf_ticker=? AND composition_date=? LIMIT 1",
            (ticker, date),
        ).fetchone()
        if not comp_meta:
            raise HTTPException(status_code=404, detail=f"No composition for {ticker} on {date}")

        rows = conn.execute(
            "SELECT component_ticker, component_name, component_identifier, component_sedol, "
            "component_weight, component_shares, component_sector, component_currency "
            "FROM equity_etf_compositions WHERE etf_ticker=? AND composition_date=? "
            "ORDER BY component_weight DESC",
            (ticker, date),
        ).fetchall()

    info_d = dict(info)
    meta_d = dict(comp_meta)
    holdings_d = [dict(r) for r in rows]

    if format == "xlsx":
        return _build_xlsx(ticker, date, info_d, meta_d, holdings_d)
    if format == "csv":
        return _build_csv(ticker, date, info_d, meta_d, holdings_d)
    if format == "json":
        return _build_json(ticker, date, info_d, meta_d, holdings_d)

    raise HTTPException(status_code=400, detail=f"Unsupported format '{format}'. Supported: xlsx, csv, json")


def populate_changes_cache(conn, ticker: str, date_from: str, date_to: str):
    """Compute changes for a date pair and insert into etf_composition_changes."""
    # Skip if already cached
    existing = conn.execute(
        "SELECT 1 FROM etf_composition_changes WHERE etf_ticker=? AND date_from=? AND date_to=? LIMIT 1",
        (ticker, date_from, date_to),
    ).fetchone()
    if existing:
        return

    changes = _compute_changes_for_pair(conn, ticker, date_from, date_to)
    if not changes:
        return

    conn.executemany(
        """
        INSERT OR IGNORE INTO etf_composition_changes
            (etf_ticker, date_from, date_to, change_type, component_identifier,
             component_name, component_ticker, component_sector,
             weight_old, weight_new, weight_delta, shares_old, shares_new)
        VALUES (:etf_ticker, :date_from, :date_to, :change_type, :component_identifier,
                :component_name, :component_ticker, :component_sector,
                :weight_old, :weight_new, :weight_delta, :shares_old, :shares_new)
        """,
        changes,
    )
    conn.commit()


@router.get("/{ticker}/changes", response_model=ChangesResponse)
def get_etf_changes(ticker: str):
    ticker = ticker.upper()
    with get_db() as conn:
        info = conn.execute(
            "SELECT ticker FROM equity_etf_info WHERE ticker=?", (ticker,)
        ).fetchone()
        if not info:
            raise HTTPException(status_code=404, detail=f"ETF '{ticker}' not found")

        dates = conn.execute(
            "SELECT DISTINCT composition_date FROM equity_etf_compositions "
            "WHERE etf_ticker=? ORDER BY composition_date ASC",
            (ticker,),
        ).fetchall()
        date_list = [r["composition_date"] for r in dates]

        # Populate cache for any un-cached consecutive pairs
        for i in range(len(date_list) - 1):
            populate_changes_cache(conn, ticker, date_list[i], date_list[i + 1])

        rows = conn.execute(
            "SELECT date_from, date_to, change_type, component_identifier, "
            "component_name, component_ticker, component_sector, "
            "weight_old, weight_new, weight_delta "
            "FROM etf_composition_changes WHERE etf_ticker=? "
            "ORDER BY date_to DESC, change_type",
            (ticker,),
        ).fetchall()

    return ChangesResponse(
        ticker=ticker,
        changes=[ChangeEvent(**dict(r)) for r in rows],
    )
